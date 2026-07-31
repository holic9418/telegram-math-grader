# -*- coding: utf-8 -*-
"""
월간 출석부 핵심 모듈.
 - 한컴(한셀) xlsx 호환 로드/저장
 - 반별 요일패턴으로 새 달 출석부 생성 (학생/서식 유지, 공휴일 병합+빨강)
 - 시트 → HTML 미리보기 렌더
"""
import io, re, zipfile, datetime, calendar, unicodedata
from copy import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border
from openpyxl.cell.cell import MergedCell
import holidays as _holidays

WEEKDAY_KR = ['월', '화', '수', '목', '금', '토', '일']


def nfc(s):
    """한글 이름을 NFC로 통일한다. 맥(NFD)·윈도우(NFC)에서 입력된 같은 이름이
    눈엔 똑같아도 코드포인트가 달라 매칭에 실패하는 걸 막는다."""
    return unicodedata.normalize('NFC', s) if isinstance(s, str) else s
KR_WD_TO_IDX = {k: i for i, k in enumerate(WEEKDAY_KR)}

# 국경일이지만 실제 휴무가 아닌 날 (수업 정상 진행) → 공휴일에서 제외
NON_OFF_HOLIDAYS = {'제헌절'}


# ── 1. 한컴 호환 로드/저장 ──────────────────────────────────────
def _normalize_hancom_bytes(data: bytes) -> bytes:
    """styles.xml 의 mc:AlternateContent 를 Fallback 내용으로 치환해 openpyxl 이 읽게 함."""
    zin = zipfile.ZipFile(io.BytesIO(data))
    styles = zin.read('xl/styles.xml').decode('utf-8')

    def repl(m):
        fb = re.search(r'<mc:Fallback>(.*?)</mc:Fallback>', m.group(0), re.S)
        return fb.group(1) if fb else ''

    styles2 = re.sub(r'<mc:AlternateContent\b.*?</mc:AlternateContent>', repl,
                     styles, flags=re.S)
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            d = zin.read(item.filename)
            if item.filename == 'xl/styles.xml':
                d = styles2.encode('utf-8')
            zout.writestr(item, d)
    zin.close()
    return out.getvalue()


def load_workbook(path_or_bytes):
    """한컴 xlsx 를 openpyxl Workbook 으로 로드 (경로 또는 bytes)."""
    if isinstance(path_or_bytes, (bytes, bytearray)):
        data = bytes(path_or_bytes)
    else:
        with open(path_or_bytes, 'rb') as f:
            data = f.read()
    try:
        return openpyxl.load_workbook(io.BytesIO(data))
    except Exception:
        return openpyxl.load_workbook(io.BytesIO(_normalize_hancom_bytes(data)))


# ── 2. 시트 구조 분석 ──────────────────────────────────────────
BLOCK_LABELS = ['출석', '수업내용', '과제수행', '다음과제', '비고']
BLOCK_SIZE = len(BLOCK_LABELS)


_DATE_AT_START = re.compile(r'^\s*(\d{1,2})/(\d{1,2})')


def date_key(v):
    """셀/문자열에서 'M/D' 추출. '7/15(화)','7/15' → '7/15'. 아니면 None.
    (날짜 옆 요일 표기가 있어도 인식하기 위한 정규화 키)"""
    if v is None:
        return None
    m = _DATE_AT_START.match(str(v).strip())
    return f"{int(m.group(1))}/{int(m.group(2))}" if m else None


def date_label(md, year=None):
    """'7/15' → '7/15(화)'. 요일을 계산해 붙인다."""
    k = date_key(md)
    if not k:
        return str(md)
    mm, dd = map(int, k.split('/'))
    try:
        wd = WEEKDAY_KR[datetime.date(year or datetime.date.today().year, mm, dd).weekday()]
        return f"{mm}/{dd}({wd})"
    except ValueError:
        return k


def _find_start_row(ws):
    """col A 에 날짜가 처음 등장하는 행 = 첫 블록 시작행."""
    for r in range(1, ws.max_row + 1):
        if date_key(ws.cell(r, 1).value):
            return r
    return None


def _last_col(ws, start):
    """헤더+첫 블록에서 실제 사용된 마지막 열."""
    last = 2
    for r in range(1, start + BLOCK_SIZE):
        for c in range(1, 60):
            if ws.cell(r, c).value not in (None, ''):
                last = max(last, c)
    for rng in ws.merged_cells.ranges:
        if rng.min_row < start + BLOCK_SIZE:
            last = max(last, rng.max_col)
    return last


def infer_schedule(ws):
    """기존 시트의 날짜들로부터 수업 요일 집합 추론."""
    start = _find_start_row(ws)
    days = set()
    year = datetime.date.today().year
    for r in range(start, ws.max_row + 1):
        k = date_key(ws.cell(r, 1).value)
        if k:
            mm, dd = k.split('/')
            try:
                days.add(datetime.date(year, int(mm), int(dd)).weekday())
            except ValueError:
                pass
    return sorted(days)


# ── 3. 공휴일 ──────────────────────────────────────────────────
def holidays_for(year):
    kr = _holidays.SouthKorea(years=[year])
    return {d: name for d, name in kr.items() if name not in NON_OFF_HOLIDAYS}


def class_days(year, month, weekday_idxs):
    n = calendar.monthrange(year, month)[1]
    out = []
    for day in range(1, n + 1):
        d = datetime.date(year, month, day)
        if d.weekday() in weekday_idxs:
            out.append(d)
    return out


# ── 4. 새 달 생성 ──────────────────────────────────────────────
STUDENT_FIRST_COL = 3   # A=날짜, B=라벨, C부터 학생/CLASS
# 반별 한 칸으로 병합할 라벨(상대행) — 수업내용, 다음과제
MERGE_LABEL_ROWS = [1, 3]


def _detect_groups(ws, start, last_col):
    """헤더의 CLASS 병합으로 학생 열 그룹 감지. 없으면 전체를 한 그룹."""
    groups = []
    for rng in ws.merged_cells.ranges:
        if rng.max_row < start and rng.min_col >= STUDENT_FIRST_COL and rng.max_col > rng.min_col:
            groups.append((rng.min_col, min(rng.max_col, last_col)))
    if groups:
        groups.sort()
        return groups
    return [(STUDENT_FIRST_COL, last_col)]


def _capture_template(ws, start, last_col):
    """첫 블록의 셀 스타일/행높이만 캡처 (병합은 재구성)."""
    styles, heights = [], []
    for i in range(BLOCK_SIZE):
        r = start + i
        styles.append({c: copy(ws.cell(r, c)._style) for c in range(1, last_col + 1)})
        heights.append(ws.row_dimensions[r].height)
    label_font = copy(ws.cell(start, 2).font)
    groups = _detect_groups(ws, start, last_col)
    return {'styles': styles, 'heights': heights, 'groups': groups,
            'last_col': last_col, 'label_font': label_font}


def regenerate_month_sheet(ws, year, month, weekday_idxs, hol=None):
    """한 시트의 날짜 블록을 (year,month) 수업요일에 맞게 새로 채운다.
       학생 열/헤더/서식은 유지하고 값만 비운다. 공휴일은 병합+빨강."""
    if hol is None:
        hol = holidays_for(year)
    start = _find_start_row(ws)
    if start is None:
        return
    last_col = _last_col(ws, start)
    tpl = _capture_template(ws, start, last_col)
    days = class_days(year, month, weekday_idxs)

    # 1) 블록 영역(>= start)의 기존 병합 모두 해제
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= start:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                # openpyxl이 미실체화 셀 병합을 못 지우는 버그 → 강제 정상화
                try:
                    ws.merged_cells.ranges.discard(rng)
                except Exception:
                    try:
                        ws.merged_cells.ranges.remove(rng)
                    except Exception:
                        pass
                for rr in range(rng.min_row, rng.max_row + 1):
                    for cc in range(rng.min_col, rng.max_col + 1):
                        if isinstance(ws._cells.get((rr, cc)), MergedCell):
                            del ws._cells[(rr, cc)]

    # 2) 블록 영역 값 비우기 (넉넉히)
    max_clear = start + BLOCK_SIZE * (len(days) + 4)
    for r in range(start, max(max_clear, ws.max_row) + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # 3) 새 블록 기록
    for i, d in enumerate(days):
        top = start + BLOCK_SIZE * i
        for k in range(BLOCK_SIZE):
            r = top + k
            for c in range(1, last_col + 1):
                if isinstance(ws.cell(r, c), MergedCell):
                    continue
                ws.cell(r, c)._style = copy(tpl['styles'][k][c])
                _strip_diagonal(ws.cell(r, c))
            if tpl['heights'][k] is not None:
                ws.row_dimensions[r].height = tpl['heights'][k]
            ws.cell(r, 2).value = BLOCK_LABELS[k]
        ws.cell(top, 1).value = f'{d.month}/{d.day}({WEEKDAY_KR[d.weekday()]})'
        ws.merge_cells(start_row=top, start_column=1, end_row=top + BLOCK_SIZE - 1, end_column=1)
        if d in hol:
            set_holiday_block(ws, top, last_col, hol[d], tpl['label_font'])
        else:
            for rel in MERGE_LABEL_ROWS:
                for (g0, g1) in tpl['groups']:
                    if g1 > g0:
                        ws.merge_cells(start_row=top + rel, start_column=g0,
                                       end_row=top + rel, end_column=g1)

    # 4) 마지막 블록 아래 잔여 영역 비우기
    last_end = start + BLOCK_SIZE * len(days) - 1
    for r in range(last_end + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = None
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.border = Border()


def generate_month(src_workbook, year, month, schedules=None):
    """src_workbook(Workbook) 을 바탕으로 새 달 Workbook 을 만들어 반환.
       schedules: {sheet_name: [weekday_idx,...]} 없으면 기존 파일에서 추론."""
    wb = src_workbook
    hol = holidays_for(year)
    schedules = schedules or {}
    for name in wb.sheetnames:
        ws = wb[name]
        wds = schedules.get(name) or infer_schedule(ws)
        regenerate_month_sheet(ws, year, month, wds, hol)
    return wb


# ── 4b. 출석 기록/조회 ─────────────────────────────────────────
def get_roster(ws):
    """{학생이름: 열번호} 반환 (헤더의 학생이름 행 기준)."""
    start = _find_start_row(ws)
    if start is None:
        return {}
    last_col = _last_col(ws, start)
    header_row = start - 1
    roster = {}
    for c in range(STUDENT_FIRST_COL, last_col + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip() and str(v).strip() != '학생이름':
            roster[nfc(str(v).strip())] = c
    return roster


def find_date_block(ws, date_str):
    """'8/4'(또는 '8/4(화)') 같은 날짜의 블록 시작행 반환 (없으면 None)."""
    start = _find_start_row(ws)
    target = date_key(date_str)
    if target is None or start is None:
        return None
    for r in range(start, ws.max_row + 1):
        if date_key(ws.cell(r, 1).value) == target:
            return r
    return None


def block_has_data(ws, top, last_col):
    """블록(5행)에 실제 기록이 있는지. 날짜·라벨(1·2열)은 제외하고 학생 열만 본다."""
    for k in range(BLOCK_SIZE):
        for c in range(STUDENT_FIRST_COL, last_col + 1):
            if ws.cell(top + k, c).value not in (None, ''):
                return True
    return False


def find_stray_blocks(ws, month=None):
    """명백히 잘못된 날짜 블록만 찾는다.
       - '중복'   : 같은 날짜가 두 번 이상 (첫 번째를 정상으로 보고 이후를 이상으로)
       - '다른 달' : 파일의 달과 다른 달의 날짜
    수업요일과 대조하지는 않는다 — 월중 시간표 변경이 정상 운영이라 오탐이 난다.
    반환: [{'top', 'date', 'reason', 'has_data'}] (행 순서)"""
    start = _find_start_row(ws)
    if start is None:
        return []
    last_col = _last_col(ws, start)
    seen, out = set(), []
    for r in range(start, ws.max_row + 1):
        ds = date_key(ws.cell(r, 1).value)
        if not ds:
            continue
        reason = None
        if ds in seen:
            reason = '중복'
        elif month is not None and int(ds.split('/')[0]) != month:
            reason = '다른 달'
        seen.add(ds)
        if reason:
            out.append({'top': r, 'date': ds, 'reason': reason,
                        'has_data': block_has_data(ws, r, last_col)})
    return out


def remove_block(ws, top):
    """블록(5행)을 비운다 — 값·테두리·행높이·병합 제거.
    호출 전에 block_has_data 로 비어 있음을 반드시 확인할 것."""
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    bottom = top + BLOCK_SIZE - 1
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= top and rng.max_row <= bottom:
            ws.unmerge_cells(str(rng))
    for r in range(top, bottom + 1):
        ws.row_dimensions[r].height = None
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.border = Border()


def attendance_recorded(ws, date_str):
    """해당 날짜의 출석 행에 기록이 있는지 확인한다.
    반환: True  = 이미 출석 입력됨
          False = 그 날짜 블록은 있으나 출석이 비어 있음(입력 필요)
          None  = 해당 날짜 블록 없음 / 공휴일·휴강 블록(확인 불가·불필요)."""
    top = find_date_block(ws, date_str)
    if top is None:
        return None
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    if _is_holiday_block(ws, top, last_col):
        return None
    att_row = top + ROW_OFFSET['출석']
    for c in get_roster(ws).values():
        if ws.cell(att_row, c).value not in (None, ''):
            return True
    return False


def rosters_summary(wb):
    """{sheet: [학생이름,...]} — 봇 파싱용."""
    out = {}
    for name in wb.sheetnames:
        out[name] = list(get_roster(wb[name]).keys())
    return out


ROW_OFFSET = {'출석': 0, '수업내용': 1, '과제수행': 2, '다음과제': 3, '비고': 4}

# 반별 수업 요일 기본값 (요일 인덱스: 월0~일6)
DEFAULT_SCHEDULES = {
    '초3': [1, 2, 3], '초4': [0, 2, 4], '초5A': [1, 2, 3], '초5B': [1, 2, 3],
    '초6A': [0, 2, 4], '초6B': [0, 2, 4], '중1AB': [0, 2, 4], '중1C': [0, 2, 4],
    '중1보충': [0, 4], '중2': [1, 2, 3], '중3': [0, 2, 4],
    '고1': [1, 3, 5], '고2': [1, 3, 5], '고3': [0, 2, 4],
}

# 반별 '출석 입력 확인' 시각 = 수업 종료 15분 뒤. {반: {요일idx(str): 'HH:MM'}}
# 요일 idx: 월0 화1 수2 목3 금4 토5 일6. 이 표에 있는 요일·시각에만 알림이 울린다.
DEFAULT_CLASS_TIMES = {
    '초3': {'1': '17:15', '2': '15:15', '3': '15:15'},          # 화 4~5시 / 수목 2~3시
    '초4': {'0': '15:15', '2': '15:15', '4': '15:15'},          # 월수금 2~3시
    '초5A': {'1': '16:15', '2': '16:15', '3': '16:15'},         # 화수목 3~4시
    '초5B': {'1': '16:15', '2': '16:15', '3': '16:15'},         # 화수목 3~4시
    '초6A': {'0': '16:15', '2': '16:15', '4': '16:15'},         # 월수금 3~4시
    '초6B': {'0': '16:15', '2': '16:15', '4': '16:15'},         # 월수금 3~4시
    '중1AB': {'0': '20:15', '4': '20:15', '2': '19:15'},        # 월금 6~8시 / 수 6~7시
    '중1C': {'0': '20:15', '4': '20:15', '2': '19:15'},         # 월금 6~8시 / 수 6~7시
    '중1보충': {'0': '18:15', '4': '18:15'},                    # 월금 5~6시
    '중2': {'1': '18:15', '3': '18:15', '2': '17:15'},          # 화목 4~6시 / 수 4~5시
    '중3': {'0': '18:15', '4': '18:15', '2': '18:15'},          # 월금 4~6시 / 수 5~6시
    '고1': {'1': '22:15', '3': '22:15', '5': '12:15'},          # 화목 8~10시 / 토 10~12시
    '고2': {'1': '20:15', '3': '20:15', '5': '14:15'},          # 화목 6~8시 / 토 12~2시
    '고3': {'0': '22:15', '2': '22:15', '4': '22:15'},          # 월수금 8~10시
}


def sheet_dates(wb):
    """{sheet: ['8/4', ...]} — 각 시트의 수업일."""
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        start = _find_start_row(ws)
        ds = []
        if start:
            for r in range(start, ws.max_row + 1):
                k = date_key(ws.cell(r, 1).value)
                if k:
                    ds.append(k)
        out[name] = ds
    return out


def next_year_month(year, month):
    return (year + 1, 1) if month == 12 else (year, month + 1)


def add_weekday_labels(wb, year=None):
    """모든 시트의 날짜 셀에 요일을 붙인다 ('7/15' → '7/15(화)').
    이미 요일이 있으면 그대로 둔다. 데이터·서식은 건드리지 않음. 반환: 바꾼 개수."""
    year = year or datetime.date.today().year
    n = 0
    for name in wb.sheetnames:
        ws = wb[name]
        start = _find_start_row(ws)
        if not start:
            continue
        for r in range(start, ws.max_row + 1):
            v = ws.cell(r, 1).value
            k = date_key(v)
            if not k or '(' in str(v):     # 날짜 아님/이미 요일표기 → 건너뜀
                continue
            ws.cell(r, 1).value = date_label(k, year)
            n += 1
    return n


def _strip_diagonal(cell):
    """셀의 대각선(사선) 테두리만 제거하고 나머지 테두리는 유지."""
    b = cell.border
    if (b.diagonal and getattr(b.diagonal, 'style', None)) or b.diagonalUp or b.diagonalDown:
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=b.bottom)


def _is_holiday_block(ws, top, last_col):
    """블록에 '글자가 있는' 사각형(여러행×여러열) 병합이 있으면 공휴일/휴강 블록으로 간주.
    빈(None)·'None' 사각형 병합은 휴강이 아니라 정리 대상(소수인원 반 등)."""
    for rng in ws.merged_cells.ranges:
        if (rng.min_row >= top and rng.max_row <= top + BLOCK_SIZE - 1
                and rng.min_col >= STUDENT_FIRST_COL
                and rng.max_row > rng.min_row and rng.max_col > rng.min_col):
            anchor = ws.cell(rng.min_row, rng.min_col).value
            if anchor not in (None, "") and str(anchor).strip().lower() != "none":
                return True
    return False


def normalize_block(ws, top, last_col, groups):
    """블록을 '전원 정상출석' 기준 깨끗한 서식으로 정리한다.
       - 학생 세로병합(결석 표시) 해제, 병합됐던 칸 테두리를 정상 칸 기준으로 통일
       - 대각선(사선) 제거
       - 수업내용/다음과제 행을 반(그룹)별 가로 한 칸으로 재병합
       공휴일/휴강 블록(사각형 병합)은 건드리지 않는다."""
    if _is_holiday_block(ws, top, last_col):
        return
    vcols = set()
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_row >= top and rng.max_row <= top + BLOCK_SIZE - 1
                and rng.min_col >= STUDENT_FIRST_COL):
            if rng.max_row > rng.min_row and rng.min_col == rng.max_col:
                vcols.add(rng.min_col)
            try:
                ws.unmerge_cells(str(rng))
            except (KeyError, ValueError):
                # openpyxl 내부 _cells 불일치로 실패하면 범위만 강제 제거
                ws.merged_cells.ranges.discard(rng)
    # 병합 해제된 학생열을 같은 행의 정상 참조열 스타일로 통일
    for r in range(top, top + BLOCK_SIZE):
        ref = next((c for c in range(STUDENT_FIRST_COL, last_col + 1) if c not in vcols), None)
        if ref is not None:
            for c in vcols:
                ws.cell(r, c)._style = copy(ws.cell(r, ref)._style)
    # 남은 대각선 제거
    for r in range(top, top + BLOCK_SIZE):
        for c in range(STUDENT_FIRST_COL, last_col + 1):
            _strip_diagonal(ws.cell(r, c))
    # 수업내용/다음과제 가로 병합 복원
    for rel in MERGE_LABEL_ROWS:
        for (g0, g1) in groups:
            if g1 > g0:
                ws.merge_cells(start_row=top + rel, start_column=g0,
                               end_row=top + rel, end_column=g1)


COLOR_RED = 'FFFF0000'
COLOR_BLUE = 'FF0000FF'
COLOR_BLACK = 'FF000000'


def set_holiday_block(ws, top, last_col, text, label_font=None):
    """블록을 공휴일/휴강 모양으로 만든다 — 학생 영역 전체를 한 칸으로 병합하고
    빨간 굵은 글씨로 text. 그 날의 기존 입력은 지워진다(전원 휴강이므로 O/X 불필요).
    label_font: 기준 글꼴(없으면 라벨 칸에서 읽음)."""
    bottom = top + BLOCK_SIZE - 1
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_row >= top and rng.max_row <= bottom
                and rng.min_col >= STUDENT_FIRST_COL):
            ws.unmerge_cells(str(rng))
    for r in range(top, bottom + 1):
        for c in range(STUDENT_FIRST_COL, last_col + 1):
            ws.cell(r, c).value = None
            _strip_diagonal(ws.cell(r, c))
    ws.merge_cells(start_row=top, start_column=STUDENT_FIRST_COL,
                   end_row=bottom, end_column=last_col)
    cell = ws.cell(top, STUDENT_FIRST_COL)
    cell.value = text
    base = label_font or ws.cell(top, 2).font
    cell.font = Font(name=base.name, size=base.sz, bold=True, color=COLOR_RED)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def clear_holiday_block(ws, top, last_col):
    """휴강/공휴일 블록을 보통 빈 블록으로 되돌린다.
    set_holiday_block 이 서식을 바꾸는 건 왼쪽 위 한 칸(빨간 굵은 글꼴)뿐이므로
    그 칸만 같은 행의 옆 칸 서식으로 되돌리고, 병합 해제·테두리 통일·가로병합
    복원은 normalize_block 에 맡긴다. 다른 블록의 서식을 베끼면 그 블록의 결석
    표시(빨강)까지 따라오므로 그렇게 하지 않는다.
    지워진 출석 기록까지 되살리지는 못한다."""
    start = _find_start_row(ws)
    if start is None or not _is_holiday_block(ws, top, last_col):
        return False
    if last_col <= STUDENT_FIRST_COL:
        return False  # 기준 삼을 옆 칸이 없음
    bottom = top + BLOCK_SIZE - 1
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_row >= top and rng.max_row <= bottom
                and rng.min_col >= STUDENT_FIRST_COL):
            ws.unmerge_cells(str(rng))
    for k in range(BLOCK_SIZE):
        for c in range(STUDENT_FIRST_COL, last_col + 1):
            ws.cell(top + k, c).value = None
        ws.cell(top + k, STUDENT_FIRST_COL)._style = copy(
            ws.cell(top + k, STUDENT_FIRST_COL + 1)._style)
    normalize_block(ws, top, last_col, _detect_groups(ws, start, last_col))
    return True


def _is_late(value):
    """지각 값인지: '지각' 이 있거나 'N분'(지각시간) 형식이면 지각."""
    s = str(value)
    return '지각' in s or bool(re.search(r'\d+\s*분', s))


def _attendance_color(value):
    """출석 값 → 글자색. 결석=빨강, 지각·조퇴=파랑, 정상=검정."""
    s = str(value)
    if _is_late(s) or '조퇴' in s:
        return COLOR_BLUE
    if '결석' in s or '결' == s.strip() or 'X' in s.upper():
        return COLOR_RED
    return COLOR_BLACK


def is_absent(value):
    """출석 값이 '결석'인지 판단한다 (지각·조퇴는 수업 참석으로 보아 제외).
    결석이면 그 학생의 과제수행·비고 칸을 비운다."""
    s = str(value)
    if _is_late(s) or '조퇴' in s:
        return False
    return 'X' in s.upper()


def _parse_md(s, year):
    """'M/D' → date. 실패 시 None."""
    try:
        m, d = map(int, str(s).split('/'))
        return datetime.date(year, m, d)
    except Exception:
        return None


def is_enrolled_during(info, monday, sunday, year):
    """info={'from':'M/D'|None,'to':'M/D'|None} 학생이 [monday,sunday] 주에 재적 중인지.
    - 그 주 시작 전에 이미 나간(to < monday) 학생은 제외 (→ 다음 주부터 안 나옴)
    - 그 주 이후에 들어오는(from > sunday) 학생도 제외"""
    if not info:
        return True
    to = _parse_md(info.get('to'), year)
    if to and to < monday:
        return False
    frm = _parse_md(info.get('from'), year)
    if frm and frm > sunday:
        return False
    return True


def _homework_color(value):
    """과제수행 값 → 글자색. 안함=빨강, 50%/절반=파랑, 완료=검정."""
    s = str(value).strip()
    if any(k in s for k in ('50', '％', '%', '△', '절반')) or s == '반':
        return COLOR_BLUE
    if any(k in s for k in ('X', 'x', '안', '미', '못', '노')):
        return COLOR_RED
    return COLOR_BLACK


def _apply_font_color(cell, argb):
    f = cell.font
    cell.font = Font(name=f.name, size=f.sz, bold=f.bold, italic=f.italic,
                     underline=f.underline, color=argb)


# ── 학적 변동 (신규등록/퇴원/담당변경) ─────────────────────────
# 출석칸에 찍는 표시
LIFECYCLE_MARK = {'신규등록': '신규등록', '전입': '담당변경', '퇴원': '퇴원', '전출': '담당변경'}
LIFECYCLE_ADD = {'신규등록', '전입'}   # 명단 추가 + 그 날짜부터 재적
LIFECYCLE_END = {'퇴원', '전출'}       # 그 날짜까지만 재적(이후 빈칸)


def _date_key(date_str):
    try:
        m, d = str(date_str).split('/')
        return (int(m), int(d))
    except Exception:
        return None


def student_active(enroll, student, date_str):
    """enroll: {student: {'from':'M/D'|None,'to':'M/D'|None}}.
    해당 날짜에 이 학생이 재적(입력 대상)인지."""
    info = (enroll or {}).get(student)
    if info is None:   # 저장은 NFC지만 파일에 NFD로 남은 옛 키도 매칭
        info = {nfc(k): v for k, v in (enroll or {}).items()}.get(nfc(student))
    if not info:
        return True
    dk = _date_key(date_str)
    if dk is None:
        return True
    fk = _date_key(info.get('from')) if info.get('from') else None
    tk = _date_key(info.get('to')) if info.get('to') else None
    if fk and dk < fk:
        return False
    if tk and dk > tk:
        return False
    return True


def add_student(ws, name):
    """시트 명단 끝에 학생을 새 열로 추가(이웃 학생열 서식 복사). 반환: 새 열번호."""
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    new_col = last_col + 1
    for r in range(1, ws.max_row + 1):
        src = ws.cell(r, last_col)
        if src.has_style:
            ws.cell(r, new_col)._style = copy(src._style)
    from openpyxl.utils import get_column_letter
    sl, dl = get_column_letter(last_col), get_column_letter(new_col)
    if sl in ws.column_dimensions and ws.column_dimensions[sl].width:
        ws.column_dimensions[dl].width = ws.column_dimensions[sl].width
    for r in range(start, ws.max_row + 1):      # 기존 블록 데이터는 비움
        ws.cell(r, new_col).value = None
    ws.cell(start - 1, new_col).value = nfc(name)    # 헤더에 이름(NFC로 저장)
    return new_col


def create_class_sheet(wb, name, roster, weekday_idxs, year, month, template=None):
    """새 반 시트를 만든다. 기존 시트를 복제해 명단·수업요일에 맞게 정리.
       - template: 구조 복제용 기준 시트(없으면 첫 시트)
       - roster: 학생 이름 목록 (빈 리스트면 명단 없는 빈 시트)
       반환: 생성된 시트 이름. 이미 있으면 ValueError."""
    if name in wb.sheetnames:
        raise ValueError(f"이미 있는 반이에요: {name}")
    if not wb.sheetnames:
        raise ValueError("구조를 복제할 기준 시트가 없어요.")
    tmpl = template if template in wb.sheetnames else wb.sheetnames[0]
    new = wb.copy_worksheet(wb[tmpl])
    new.title = name
    if roster:
        set_roster(wb, name, roster)
    regenerate_month_sheet(wb[name], year, month, weekday_idxs)
    return name


def set_roster(wb, sheet, target_names):
    """시트의 학생 명단을 target_names(순서대로)로 맞춘다.
       - 열 수를 늘리거나(끝 학생 서식 복사) 줄이고, 헤더에 이름을 채운다.
       - 블록 날짜/값은 건드리지 않으므로, 새 시트는 이후 regenerate_month_sheet 로 정리."""
    ws = wb[sheet]
    start = _find_start_row(ws)
    header = start - 1
    need = len(target_names)
    cur = sorted(get_roster(ws).items(), key=lambda x: x[1])
    while len(cur) < need:                       # 부족 → 끝에 학생 열 추가
        add_student(ws, f"__n{len(cur)}")
        cur = sorted(get_roster(ws).items(), key=lambda x: x[1])
    if len(cur) > need:                          # 남음 → 뒤쪽 열 제거
        rebuild_without_students(wb, sheet, [nm for nm, _ in cur[need:]])
        ws = wb[sheet]
    # 헤더 이름을 target 으로 (열은 3부터 연속). NFC로 저장해 매칭 어긋남 방지.
    for i, nm in enumerate(target_names):
        ws.cell(header, STUDENT_FIRST_COL + i).value = nfc(nm)


def rebuild_without_students(wb, sheet, drop_names):
    """sheet 에서 drop_names 학생 열을 제거한 새 시트로 교체(값·서식·병합 유지).
    delete_cols가 병합을 깨뜨리므로 필요한 열만 복사해 재구성한다."""
    from openpyxl.utils import get_column_letter
    if not drop_names:
        return
    ws = wb[sheet]
    roster = get_roster(ws)
    drop_cols = {roster[n] for n in drop_names if n in roster}
    if not drop_cols:
        return
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    keep = [c for c in range(STUDENT_FIRST_COL, last_col + 1) if c not in drop_cols]
    colmap = {1: 1, 2: 2}
    for i, sc in enumerate(keep):
        colmap[sc] = STUDENT_FIRST_COL + i
    idx = wb.sheetnames.index(sheet)
    dest = wb.create_sheet(sheet + "__tmp")
    max_row = ws.max_row
    for sc, dc in colmap.items():
        for r in range(1, max_row + 1):
            s = ws.cell(r, sc)
            dest.cell(r, dc).value = s.value
            if s.has_style:
                dest.cell(r, dc)._style = copy(s._style)
        sl, dl = get_column_letter(sc), get_column_letter(dc)
        if sl in ws.column_dimensions and ws.column_dimensions[sl].width:
            dest.column_dimensions[dl].width = ws.column_dimensions[sl].width
    for r in range(1, max_row + 1):
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None:
            dest.row_dimensions[r].height = ws.row_dimensions[r].height
    for rng in list(ws.merged_cells.ranges):
        kept = [colmap[c] for c in range(rng.min_col, rng.max_col + 1) if c in colmap]
        if not kept:
            continue  # 병합이 전부 삭제열 안이면 버림
        r0, r1, c0, c1 = rng.min_row, rng.max_row, min(kept), max(kept)
        if r1 > r0 or c1 > c0:  # 단일 셀로 줄면 병합 불필요
            dest.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
    del wb[sheet]
    dest.title = sheet
    wb.move_sheet(sheet, offset=idx - wb.sheetnames.index(sheet))


def _write_value(ws, r, c, val):
    """(r,c) 셀에 값을 쓴다. 병합된 셀(읽기 전용)이면 그 병합을 먼저 풀고 쓴다.
    소수 인원 반처럼 블록 전체가 한 칸으로 병합된 경우에도 안전하게 기록."""
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    ws.merged_cells.ranges.discard(rng)
                break
        cell = ws.cell(r, c)
    cell.value = val


def write_attendance(wb, sheet, date_str, data, enroll=None):
    """data 예:
       {'출석': {'김규림':'O','남우현':'X(결석)'},
        '수업내용': '분수의 나눗셈',
        '과제수행': {'김규림':'O'},
        '다음과제': '42쪽',
        '비고': {'남우현':'보강필요'} 또는 '단체공지'}
       반환: (기록 목록, 경고 목록)"""
    if sheet not in wb.sheetnames:
        raise KeyError(f'시트 없음: {sheet}')
    ws = wb[sheet]
    top = find_date_block(ws, date_str)
    if top is None:
        raise ValueError(f"{sheet}에 '{date_str}' 날짜가 없습니다")
    # 기록 전에 이 블록의 서식을 깨끗하게 정리(결석 세로병합·대각선 제거 등)
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    groups = _detect_groups(ws, start, last_col)
    normalize_block(ws, top, last_col, groups)
    roster = get_roster(ws)
    written, warnings = [], []

    def _norm_keys(block):
        """이름을 키로 쓰는 블록의 키를 NFC로 통일(맥 NFD 입력 대비). 문자열이면 그대로."""
        return {nfc(k): v for k, v in block.items()} if isinstance(block, dict) else block

    # 학적 변동 처리: 명단 추가(신규·전입) + 출석칸에 표시
    life = _norm_keys(data.get('학적') or {})
    for st, event in life.items():
        if event in LIFECYCLE_ADD and st not in roster:
            roster[st] = add_student(ws, st)
            last_col = _last_col(ws, start)
            groups = _detect_groups(ws, start, last_col)
        mark = LIFECYCLE_MARK.get(event)
        if not mark:
            continue
        if st not in roster:
            warnings.append(f"학적: '{st}' 학생을 못 찾음")
            continue
        cell = ws.cell(top + ROW_OFFSET['출석'], roster[st])
        if isinstance(cell, MergedCell):
            warnings.append(f"학적: '{st}' 칸이 병합돼 있어 건너뜀")
            continue
        cell.value = mark
        _apply_font_color(cell, _attendance_color(mark))
        written.append(f"학적 · {st} → {mark}")

    # 결석(지각·조퇴 제외) 학생: 과제수행·비고 칸을 비운다
    absent = {
        st for st, val in _norm_keys(data.get('출석') or {}).items()
        if st in roster and st not in life and is_absent(val)
    }
    for st in absent:
        for label in ('과제수행', '비고'):
            cell = ws.cell(top + ROW_OFFSET[label], roster[st])
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            _apply_font_color(cell, COLOR_BLACK)

    def put(label, student, value):
        r = top + ROW_OFFSET[label]
        if student not in roster:
            warnings.append(f"{label}: '{student}' 학생을 못 찾음")
            return
        c = roster[student]
        cell = ws.cell(r, c)
        if isinstance(cell, MergedCell):
            rng = next((rr for rr in list(ws.merged_cells.ranges)
                        if rr.min_row <= r <= rr.max_row and rr.min_col <= c <= rr.max_col), None)
            # 블록 안에 갇힌 병합(소수인원 반 등)은 풀고 기록.
            # 전출처럼 여러 블록에 걸친 세로 병합은 그대로 두고 건너뜀.
            if rng is not None and rng.min_row >= top and rng.max_row <= top + BLOCK_SIZE - 1:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    ws.merged_cells.ranges.discard(rng)
                cell = ws.cell(r, c)
            else:
                warnings.append(f"{label}: '{student}' 칸이 병합돼 있어 건너뜀(전출 등)")
                return
        cell.value = value
        if label == '출석':
            _apply_font_color(cell, _attendance_color(value))
        elif label == '과제수행':
            _apply_font_color(cell, _homework_color(value))
        written.append(f"{label} · {student} = {value}")

    for label in ('출석', '과제수행', '비고'):
        block = _norm_keys(data.get(label))
        if isinstance(block, dict):
            for st, val in block.items():
                if st in life:
                    continue  # 학적 표시로 이미 처리한 학생
                if label in ('과제수행', '비고') and st in absent:
                    continue  # 결석 학생의 과제·비고는 위에서 비웠으므로 건너뜀
                if st in roster and not student_active(enroll, st, date_str):
                    if label == '출석':
                        warnings.append(f"{st}: 재적 기간이 아니라 건너뜀")
                    continue  # 퇴원/전출 이후 또는 등록 전 → 빈칸 유지
                put(label, st, val)
        elif isinstance(block, str) and block.strip():
            # 단체 문자열 → 라벨행 첫 학생칸에 기록
            r = top + ROW_OFFSET[label]
            _write_value(ws, r, STUDENT_FIRST_COL, block)
            written.append(f"{label} = {block}")

    for label in ('수업내용', '다음과제'):
        val = data.get(label)
        if isinstance(val, str) and val.strip():
            r = top + ROW_OFFSET[label]
            _write_value(ws, r, STUDENT_FIRST_COL, val)  # 병합돼 있으면 풀고 씀
            written.append(f"{label} = {val}")

    # 비고 행 라벨 변경 (일일테스트 등) — B열 라벨만 이 블록에서 교체
    new_label = data.get('비고라벨') or data.get('비고제목')
    if isinstance(new_label, str) and new_label.strip():
        _write_value(ws, top + ROW_OFFSET['비고'], 2, new_label.strip())
        written.append(f"비고 라벨 → {new_label.strip()}")

    return written, warnings


# ── 5. HTML 미리보기 ───────────────────────────────────────────
def _argb_to_css(argb):
    if not argb or not isinstance(argb, str):
        return None
    s = argb[-6:] if len(argb) >= 6 else argb
    if s.upper() in ('000000',):
        return '#000'
    return '#' + s


def render_html(ws, max_rows=46):
    start = _find_start_row(ws)
    last_col = _last_col(ws, start)
    span = {}
    skip = set()
    for rng in ws.merged_cells.ranges:
        span[(rng.min_row, rng.min_col)] = (rng.max_row - rng.min_row + 1,
                                            rng.max_col - rng.min_col + 1)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    skip.add((r, c))
    rows_html = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        cells = []
        for c in range(1, last_col + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(r, c)
            v = cell.value if cell.value is not None else ''
            rs, cs = span.get((r, c), (1, 1))
            cls = []
            try:
                col = cell.font.color
                if col and col.rgb and isinstance(col.rgb, str) and col.rgb.lower().endswith('ff0000'):
                    cls.append('h')          # holiday red
            except Exception:
                pass
            if c == 2:
                cls.append('lbl')            # label column
            attr = f' rowspan="{rs}"' if rs > 1 else ''
            attr += f' colspan="{cs}"' if cs > 1 else ''
            attr += f' class="{" ".join(cls)}"' if cls else ''
            cells.append(f'<td{attr}>{v}</td>')
        rows_html.append('<tr>' + ''.join(cells) + '</tr>')
    return '<table class="att">' + ''.join(rows_html) + '</table>'
